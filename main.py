import csv
import openpyxl

# Create a new workbook and active sheet
wb = openpyxl.Workbook()
ws = wb.active

# Open the CSV file for reading
with open('csv_file.csv', 'r') as csv_file:
    # Create a CSV file reader
    file_reader = csv.DictReader(csv_file)

    # Fill headers in the first row of the Excel sheet
    for i in range(2, 8):
        ws.cell(row=1, column=i).value = f'Person {i - 1}'

    # Create a dictionary to store data from the CSV file
    res_dict = {'id': [], 'name': [], 'phone': []}

    # Process each row in the CSV file
    for row in file_reader:
        for key in row:
            # Check if the key is present in the res_dict dictionary
            if key in res_dict:
                res_dict[key].append(row[key])

    # Fill data from the dictionary into the Excel sheet
    for key in res_dict:
        ws.append([key] + res_dict[key])

    # Save the workbook to a file
    wb.save('excel_file.xlsx')

# Load the created Excel file
workbook = openpyxl.load_workbook('excel_file.xlsx')

# Print the contents of the Excel sheet
for row in workbook.active:
    for cell in row:
        print(str(cell.value).ljust(10), end=' ')
    print()

# Close the Excel file
workbook.close()
